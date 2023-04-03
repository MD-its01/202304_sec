import os
import re
import openpyxl
import configparser
import json

CONFFILE = 'param.ini'         # iniファイル
# ConfigParserクラスをインスタンス化
conf = configparser.ConfigParser()

# -----------------------------------------------
# 概要   :iniファイルからパラメータと名前リストを取得
# 引数   :なし
# 戻り値 :dir_path, month, nameList
# 詳細   :
#-----------------------------------------------
def getParam() :
    print('--- getParam START---')

    # INIファイル 読み込み
    conf.read(CONFFILE, encoding='utf-8')

    dir_path = conf.get('DATA', 'DIR')
    month = conf.get('DATA', 'MONTH')
    nameList = json.loads(conf.get('MEMBER', 'NAME'))

    print('dir_path:', dir_path)
    print('month:', month)
    print('nameList:', nameList)

    print('--- getParam END---')

    return dir_path, int(month), nameList

# iniファイルからパラメータを取得
in_dir, in_month, menberList = getParam()

#----------------------------------
# 定数
COL_START_CNT = 16 # 項番開始の行No
EXCEL_SAVE_TITLE = 'セキュリティ点検結果.xlsx'

data_dic = {}
fileList = []
#----------------------------------
# 指定フォルダのファイル一覧取得
for root, dirs, files in os.walk('data'):
    for file in files:
        filePath = os.path.join(root, file)
        fileList.append(filePath)
print('fileList:', fileList)

#----------------------------------
# ファイル一覧の数分　LOOP
for file in fileList:

    # ファイル単位で初期化
    data_list = []

    #----------------------------------
    # 拡張子チェック(末尾指定文字列チェック) [*.xlsx]以外はスキップ
    if not file.endswith('.xlsx'):
        # エクセル以外は処理しない
        print('エクセル以外 file:',file)
        continue

    #----------------------------------
    # ファイル指定して,Workbookオブジェクトを取得
    wb = openpyxl.load_workbook(file)

    #----------------------------------
    # シート名確認
    print('wb.sheetnames:', wb.sheetnames)
    sheet_name = wb.sheetnames[0]

    #----------------------------------
    # Worksheetオブジェクトを取得
    sheet = wb[sheet_name]

    #----------------------------------
    # 基本情報取得 社員番号,点検者,確認者
    # 社員番号の取得
    res_num = re.sub(r'\D', '', sheet['c4'].value)
    shain_no = int(res_num)
    print('shain_no:', shain_no)
    data_list.append(shain_no)

    # 点検者名
    tenkensha = sheet['c5'].value
    print('tenkensha:', tenkensha)
    data_list.append(tenkensha)

    # 存在チェック
    if (''.join(tenkensha.split()) in menberList) :
        # 存在するメンバーはリストから削除
        menberList.remove(''.join(tenkensha.split()))

    # 確認者
    kakuninsha =sheet['c6'].value
    print('kakuninsha:', kakuninsha)
    data_list.append(kakuninsha)

    # 項番初期値
    col_cnt = COL_START_CNT

    #----------------------------------
    # 項番の最大値を取得する為、空白行になるまで LOOP
    while True :

        koban = 'a' + str(col_cnt)
        no_val = sheet[koban].value

        # 値が存在しないセルの場合 LOOP終了
        if not no_val :
            break

        col_cnt = col_cnt + 1

    print('■ END LOOP no_val:', no_val)
    print('■ END LOOP col_cnt:', col_cnt)

    # 対象行数,終了行No
    count = col_cnt - COL_START_CNT
    col_end_cnt = col_cnt                      #-1

    print('count:', count)
    print('col_end_cnt:', col_end_cnt)

    #----------------------------------
    # 月を取得 
    # 6ヶ月[e15:j15]固定で実施
    for row in sheet['e15:j15'] :
        for cell in row :
            print('月：cell:', cell)
            print('月：cell.value:', cell.value)

            # datetime.datetime → strへ
            str_ymd = str(cell.value)

            # 先頭から[-]のインデックス取得
            idx = str_ymd.find('-')

            # スライスで文字列を抽出 「2桁取得」
            tmp_month = str_ymd[idx+1: idx+3]

            # 数値型へ
            month = int(tmp_month)

            # 対象月以外、次の月へ
            if in_month != month :
                print('対象月以外:', month)
                continue

            # 対象月の場合、座標取得
            cell_zahyo = cell.coordinate
            print('Excel座標 cell.coordinate:', cell.coordinate)
            print('type:', type(cell.coordinate))
            print('cell.coordinate[0]:', cell.coordinate[0])

            # 列の取得
            cell_row = cell.coordinate[0]

            # デフォルトは'○':確認者未チェック
            chk_flg = '○'
            biko = '確認者未チェック'

            for y in range(COL_START_CNT, col_end_cnt):
                # チェック対象の座標取得
                zahyo = cell_row + str(y)

                # チェック対象の値の取得
                value  = sheet[zahyo].value
                print('zahyo :', zahyo)
                print('value:', value)

                # 未設定場合
                if not value :
                    # 1つでも未設定の値があるチェックフラグを'×':点検者未入力を設定
                    chk_flg = '×'
                    biko = '点検者未入力'
                    break

            # 確認者の記入チェック
            zahyo = cell_row + str(col_end_cnt + 1)
            value  = sheet[zahyo].value
            print('確認者 zahyo :', zahyo)
            print('確認者 value:', value)

            # 確認者に値がある場合
            if chk_flg == '○' and value :
                # 1つでも未設定の値があるチェックフラグを'×':点検者未入力を設定
                chk_flg = '●'
                biko = '確認者 確認済'

            data_list.append(chk_flg)
            data_list.append(biko)

    data_dic[shain_no] = data_list
    print('data_dic:', data_dic)

print('menberList:', menberList)

cnt = 1
#----------------------------------
# ファイルが存在しないメンバーを追加
for notfileName in menberList :
    notExistNo = 9000 + cnt
    # 社員番号, 点検者名, 確認者名, チェック状況, 備考
    data_dic[notExistNo] = [notExistNo, notfileName, '', '×', 'ファイルが存在しない']
    cnt += 1

#----------------------------------
# key(社員番号)で並び変え
sort_dic = dict(sorted(data_dic.items()))
print('sort_dic:', sort_dic)

######################################################
# エクセルに出力
######################################################
#----------------------------------
# Excelファイルの新規作成
wb = openpyxl.Workbook()

#-------------------------------------------------
# デフォルトで、作成される「Sheet」シートを削除
wb.remove(wb['Sheet'])

#-------------------------------------------------
# シートの作成
sheet_title = str(in_month) + '月'
ws = wb.create_sheet(title=sheet_title)

#-------------------------------------------------
# ヘッダ設定
i = 1
ws.cell(row=i,column=1).value = '社員番号'
ws.cell(row=i,column=2).value = '点検者名'
ws.cell(row=i,column=3).value = '確認者名'
ws.cell(row=i,column=4).value = 'チェック状況'
ws.cell(row=i,column=5).value = '備考'


#-------------------------------------------------
# データ設定
for rec in sort_dic.values() :
    i += 1
    ws.cell(row=i,column=1).value = rec[0]
    ws.cell(row=i,column=2).value = rec[1]
    ws.cell(row=i,column=3).value = rec[2]
    ws.cell(row=i,column=4).value = rec[3]
    ws.cell(row=i,column=5).value = rec[4]

# -------------------------------------------------
# エクセル保存
wb.save(EXCEL_SAVE_TITLE)
